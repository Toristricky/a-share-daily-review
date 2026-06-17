import json
import math
import sys
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_FILE = DATA_DIR / "review.json"
DATA_JS_FILE = DATA_DIR / "review-data.js"
HISTORY_FILE = DATA_DIR / "history.json"
LOG_FILE = BASE_DIR / "refresh.log"
WORKBOOK_FILE = BASE_DIR / "每日A股复盘表.xlsx"

EASTMONEY_UT = "bd1d9ddb04089700cf9c27f6f7426281"
USE_FAST_FALLBACK = True
SINA_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Referer": "https://finance.sina.com.cn",
}

CONCEPT_RULES = [
    ("CPO/光通信", ["太辰光", "光库科技", "长进光子", "铭普光磁", "通鼎互联", "欧陆通", "长盈通", "杭电股份"]),
    ("半导体", ["南芯科技", "雅创电子", "大普微", "中巨芯", "云汉芯城", "华兴源创", "康强电子", "普冉股份", "兆易创新", "香农芯创", "飞天诚信"]),
    ("半导体设备/先进封装", ["盛美上海", "盛剑科技", "亚翔集成", "赛腾股份", "铂力特", "锐翔智能"]),
    ("PCB", ["本川智能", "逸豪新材", "胜蓝股份", "红板科技", "中京电子", "崇达技术", "沪电股份", "科翔股份", "一博科技", "深南电路", "超声电子", "鹏鼎控股", "华正新材", "宏昌电子"]),
    ("消费电子", ["精研科技", "胜蓝股份", "雅创电子", "达瑞电子", "长信科技", "隆扬电子", "光弘科技", "美迪凯", "昀冢科技", "智微智能", "联域股份"]),
    ("面板/显示", ["TCL科技", "京东方", "深天马", "彩虹股份", "沃格光电", "莱宝高科", "凯盛科技", "长信科技"]),
    ("机器人", ["德恩精工", "先导智能", "科大智能", "乔锋智能", "安达智能", "正业科技"]),
    ("智能制造", ["德恩精工", "先导智能", "科大智能", "乔锋智能", "安达智能", "通用电梯"]),
    ("工业母机", ["德恩精工", "鑫宏业", "乔锋智能", "恒锋工具"]),
    ("汽车零部件", ["肇民科技", "胜蓝股份", "瑞玛精密", "华锋股份", "星徽股份", "宏达电子", "三联锻造", "华阳集团", "永贵电器"]),
    ("铜连接/高速连接", ["胜蓝股份", "雅创电子", "本川智能", "太辰光", "鑫宏业", "逸豪新材"]),
    ("电梯", ["通用电梯", "快意电梯"]),
    ("水利/泵阀", ["南方泵业", "三川智慧", "南方泵业"]),
    ("商业零售", ["供销大集", "张小泉", "金马游乐"]),
    ("新材料/化工", ["金戈新材", "逸豪新材", "同宇新材", "国际复材", "锦华新材", "有研粉材", "有研复材", "金博股份", "瑞华泰", "泰和新材", "光华科技", "诺德股份", "昊华科技", "东阳光", "再升科技", "铜峰电子", "江海股份", "艾华集团", "海星股份", "温州宏丰", "中材科技", "中国巨石", "九鼎新材"]),
    ("建材/玻璃", ["科顺股份", "凯盛新能", "旗滨集团", "蒙娜丽莎", "上峰材料", "中国巨石", "中材科技", "凯盛科技"]),
    ("复合材料", ["国际复材", "同宇新材", "锦华新材", "金戈新材", "中国巨石", "中材科技", "九鼎新材", "泰和新材"]),
    ("锂电池", ["亿纬锂能", "珠海冠宇", "多氟多", "盛新锂能", "天山铝业", "先导智能"]),
    ("固态电池", ["亿纬锂能", "珠海冠宇", "多氟多", "先导智能"]),
    ("专精特新", ["金戈新材", "鸿仕达", "德恩精工", "南芯科技", "长进光子", "雅创电子", "本川智能", "逸豪新材"]),
    ("北交所", ["金戈新材", "鸿仕达", "新睿电子", "戈碧迦", "锦华新材"]),
    ("ST摘帽/重整", ["ST南都", "ST南新", "退市岩石", "*ST", "ST"]),
    ("大消费", ["张小泉", "供销大集", "金马游乐"]),
    ("医药/消费", ["海辰药业", "粤万年青", "正川股份", "康惠股份", "莲花控股", "爱普股份", "安德利", "大连友谊", "嘉欣丝绸", "南卫股份"]),
    ("基建工程", ["中国化学", "勘设股份", "镇海股份", "中化装备", "中核科技", "江钨装备", "长城电工", "展鹏科技"]),
    ("能源/油气", ["中远海能", "贝肯能源", "滨海能源", "冰轮环境"]),
    ("纺织检测", ["中纺标", "天纺标", "嘉欣丝绸", "欣龙控股"]),
    ("AI算力", ["云天励飞", "格灵深瞳", "华曙高科", "科大智能", "依米康"]),
    ("低空经济/航空航天", ["超卓航科", "宏达电子", "宗申动力"]),
]

CONCEPT_REASON = {
    "CPO/光通信": "AI算力需求带动高速光模块、光通信链条活跃",
    "半导体": "国产替代与AI硬件需求共振，资金关注芯片设计/设备材料",
    "半导体设备/先进封装": "半导体设备、洁净室和先进制造环节受资金追捧",
    "PCB": "AI服务器与高速连接需求改善，PCB/铜连接方向走强",
    "消费电子": "端侧AI、连接器和零部件弹性带动消费电子修复",
    "面板/显示": "面板、显示模组和光电链条放量走强",
    "机器人": "机器人产业催化持续，执行器、控制、设备方向活跃",
    "智能制造": "自动化设备和制造升级主题扩散",
    "工业母机": "高端制造设备、机床和自动化方向资金回流",
    "汽车零部件": "智能车、连接器和零部件补涨扩散",
    "铜连接/高速连接": "AI服务器高速传输链条强势，铜连接替代预期升温",
    "电梯": "设备更新和低位补涨资金关注",
    "水利/泵阀": "水利建设、设备更新和泵阀方向异动",
    "商业零售": "消费复苏预期和低价题材活跃",
    "新材料/化工": "高端制造材料、化工材料和涨价线索驱动",
    "建材/玻璃": "玻璃、建材和防水材料方向低位修复",
    "复合材料": "轻量化和高端制造材料方向受关注",
    "锂电池": "新能源产业链反弹，电池和材料方向修复",
    "固态电池": "固态电池产业化预期带动材料与设备方向",
    "专精特新": "小盘高弹性方向活跃，资金偏好细分龙头和北交所映射",
    "北交所": "北交所高弹性品种活跃，短线情绪扩散",
    "ST摘帽/重整": "重整、摘帽和低价修复预期带动",
    "大消费": "消费修复和低价消费题材活跃",
    "医药/消费": "医药、食品消费和低位消费题材活跃",
    "基建工程": "基建工程、工业装备和设备更新方向走强",
    "能源/油气": "能源运输、油气设备和传统能源方向异动",
    "纺织检测": "北交所高弹性与纺织检测细分方向活跃",
    "AI算力": "算力投资和AI应用扩散带动硬件链条",
    "低空经济/航空航天": "低空经济政策和产业化预期反复活跃",
    "其他涨停": "单股事件、低位补涨或暂未归入主线概念",
}


def log(message):
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(f"[{stamp}] {message}\n")


def fetch_json(url, params):
    query = urllib.parse.urlencode(params)
    req = urllib.request.Request(
        f"{url}?{query}",
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://quote.eastmoney.com/",
        },
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fetch_text(url, params=None, headers=None, encoding="utf-8"):
    query = urllib.parse.urlencode(params or {})
    target = f"{url}?{query}" if query else url
    req = urllib.request.Request(
        target,
        headers=headers
        or {
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://quote.eastmoney.com/",
        },
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        return resp.read().decode(encoding, errors="ignore")


def safe_num(value, default=0):
    if value in (None, "-", ""):
        return default
    try:
        number = float(value)
        if math.isnan(number):
            return default
        return number
    except Exception:
        return default


def pct(value):
    return round(safe_num(value), 2)


def get_indices():
    if USE_FAST_FALLBACK:
        try:
            return get_tencent_indices()
        except Exception as exc:
            log(f"tencent index fetch failed, falling back to sina index: {exc}")
            return get_sina_indices()
    try:
        data = fetch_json(
            "https://push2.eastmoney.com/api/qt/ulist.np/get",
            {
                "fltt": 2,
                "invt": 2,
                "fields": "f12,f14,f2,f3,f4,f6",
                "secids": "1.000001,0.399001,0.399006,1.000688",
            },
        )
        rows = data.get("data", {}).get("diff", []) or []
        return [
            {
                "code": r.get("f12"),
                "name": r.get("f14"),
                "price": safe_num(r.get("f2")),
                "change_pct": pct(r.get("f3")),
                "change": safe_num(r.get("f4")),
                "amount_yuan": safe_num(r.get("f6")),
            }
            for r in rows
        ]
    except Exception as exc:
        log(f"eastmoney index fetch failed, falling back to tencent: {exc}")

    return get_tencent_indices()


def get_tencent_indices():
    text = fetch_text(
        "http://qt.gtimg.cn/q=sh000001,sz399001,sz399006",
        headers={"User-Agent": "Mozilla/5.0"},
        encoding="gbk",
    )
    result = []
    for line in text.splitlines():
        if not line or '="' not in line:
            continue
        payload = line.split('="', 1)[1].rstrip('";')
        parts = payload.split("~")
        if len(parts) < 38:
            continue
        result.append(
            {
                "code": parts[2],
                "name": parts[1],
                "price": safe_num(parts[3]),
                "change_pct": pct(parts[32]),
                "change": safe_num(parts[31]),
                "amount_yuan": safe_num(parts[37]) * 10000,
                "quote_time": parts[30] if len(parts) > 30 else "",
                "source": "Tencent qt.gtimg.cn",
            }
        )
    return result


def get_sina_indices():
    text = fetch_text(
        "http://hq.sinajs.cn/list=sh000001,sz399001,sz399006",
        headers=SINA_HEADERS,
        encoding="gbk",
    )
    result = []
    code_map = {"sh000001": "000001", "sz399001": "399001", "sz399006": "399006"}
    for line in text.splitlines():
        if '="' not in line:
            continue
        symbol = line.split("var hq_str_", 1)[1].split("=", 1)[0]
        payload = line.split('="', 1)[1].rstrip('";')
        parts = payload.split(",")
        if len(parts) < 32:
            continue
        name = parts[0]
        open_price = safe_num(parts[1])
        pre_close = safe_num(parts[2])
        price = safe_num(parts[3])
        high = safe_num(parts[4])
        low = safe_num(parts[5])
        amount = safe_num(parts[9])
        change = price - pre_close
        change_pct = change / pre_close * 100 if pre_close else 0
        result.append(
            {
                "code": code_map.get(symbol, symbol),
                "name": name,
                "price": round(price, 2),
                "change_pct": round(change_pct, 2),
                "change": round(change, 2),
                "open": open_price,
                "high": high,
                "low": low,
                "amount_yuan": amount,
                "quote_time": f"{parts[30]} {parts[31]}" if len(parts) > 31 else "",
                "source": "Sina hq.sinajs.cn",
            }
        )
    return result


def get_a_shares():
    if USE_FAST_FALLBACK:
        return get_sina_rank_pages("changepercent", 0, max_pages=4, stop_below=4.8)
    try:
        data = fetch_json(
            "https://push2.eastmoney.com/api/qt/clist/get",
            {
                "pn": 1,
                "pz": 6000,
                "po": 1,
                "np": 1,
                "ut": EASTMONEY_UT,
                "fltt": 2,
                "invt": 2,
                "fid": "f3",
                "fs": "m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23",
                "fields": "f12,f14,f2,f3,f4,f5,f6,f8,f62",
            },
        )
        rows = data.get("data", {}).get("diff", []) or []
        return [
            {
                "code": r.get("f12"),
                "name": r.get("f14"),
                "price": safe_num(r.get("f2")),
                "change_pct": pct(r.get("f3")),
                "amount_yuan": safe_num(r.get("f6")),
                "turnover_pct": pct(r.get("f8")),
                "main_net_yuan": safe_num(r.get("f62")),
            }
            for r in rows
        ]
    except Exception as exc:
        log(f"eastmoney stock list failed, falling back to sina rank data: {exc}")
        return get_sina_rank("changepercent", 0, 180)


def get_sina_rank(sort, asc, count=80):
    raw = fetch_text(
        "http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeData",
        {
            "page": 1,
            "num": count,
            "sort": sort,
            "asc": asc,
            "node": "hs_a",
            "symbol": "",
            "_s_r_a": "page",
        },
        headers=SINA_HEADERS,
    )
    rows = json.loads(raw)
    return [
        {
            "code": r.get("code"),
            "name": r.get("name"),
            "price": safe_num(r.get("trade")),
            "change_pct": pct(r.get("changepercent")),
            "price_change": safe_num(r.get("pricechange")),
            "pre_close": safe_num(r.get("settlement")),
            "open": safe_num(r.get("open")),
            "high": safe_num(r.get("high")),
            "low": safe_num(r.get("low")),
            "volume": safe_num(r.get("volume")),
            "amount_yuan": safe_num(r.get("amount")),
            "turnover_pct": pct(r.get("turnoverratio")),
            "main_net_yuan": 0,
            "pe": safe_num(r.get("per")),
            "pb": safe_num(r.get("pb")),
            "market_cap_wan": safe_num(r.get("mktcap")),
            "float_market_cap_wan": safe_num(r.get("nmc")),
            "ticktime": r.get("ticktime") or "",
            "source": "Sina Market_Center.getHQNodeData",
        }
        for r in rows
    ]


def get_sina_rank_pages(sort, asc, max_pages=4, stop_below=None):
    stocks = []
    seen = set()
    for page in range(1, max_pages + 1):
        try:
            raw = fetch_text(
                "http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeData",
                {
                    "page": page,
                    "num": 100,
                    "sort": sort,
                    "asc": asc,
                    "node": "hs_a",
                    "symbol": "",
                    "_s_r_a": "page",
                },
                headers=SINA_HEADERS,
            )
            rows = json.loads(raw)
        except Exception as exc:
            if stocks:
                log(f"sina rank page {page} failed, using {len(stocks)} rows fetched: {exc}")
                break
            raise
        if not rows:
            break
        converted = [
            {
                "code": r.get("symbol", "")[-6:],
                "name": r.get("name", ""),
                "price": safe_num(r.get("trade")),
                "change_pct": safe_num(r.get("changepercent")),
                "price_change": safe_num(r.get("pricechange")),
                "pre_close": safe_num(r.get("settlement")),
                "open": safe_num(r.get("open")),
                "high": safe_num(r.get("high")),
                "low": safe_num(r.get("low")),
                "volume": safe_num(r.get("volume")),
                "amount_yuan": safe_num(r.get("amount")),
                "turnover_pct": pct(r.get("turnoverratio")),
                "main_net_yuan": 0,
                "pb": safe_num(r.get("pb"), None),
                "pe": safe_num(r.get("per"), None),
                "industry": r.get("industry") or "",
                "ticktime": r.get("ticktime") or "",
                "source": "Sina Market_Center.getHQNodeData",
            }
            for r in rows
        ]
        for item in converted:
            key = item["code"] or item["name"]
            if key and key not in seen:
                stocks.append(item)
                seen.add(key)
        if stop_below is not None and converted[-1]["change_pct"] < stop_below:
            break
    return stocks


def get_boards():
    if USE_FAST_FALLBACK:
        return []
    try:
        data = fetch_json(
            "https://push2.eastmoney.com/api/qt/clist/get",
            {
                "pn": 1,
                "pz": 80,
                "po": 1,
                "np": 1,
                "ut": EASTMONEY_UT,
                "fltt": 2,
                "invt": 2,
                "fid": "f3",
                "fs": "m:90+t:2,m:90+t:3",
                "fields": "f12,f14,f3,f62,f128,f140",
            },
        )
        rows = data.get("data", {}).get("diff", []) or []
        return [
            {
                "code": r.get("f12"),
                "name": r.get("f14"),
                "change_pct": pct(r.get("f3")),
                "main_net_yuan": safe_num(r.get("f62")),
                "leader": r.get("f128") or r.get("f140") or "",
            }
            for r in rows
        ]
    except Exception as exc:
        log(f"board fetch failed: {exc}")
        return []


def yuan_to_yi(value):
    return round(safe_num(value) / 100000000, 2)


def market_state(indices, up_count, down_count, limit_up, limit_down):
    avg_index = sum(i["change_pct"] for i in indices[:3]) / max(len(indices[:3]), 1)
    breadth = up_count - down_count
    if avg_index > 1 and breadth > 1000 and limit_up >= 60:
        return "强势"
    if avg_index > 0.25 and breadth > 0:
        return "震荡偏强"
    if avg_index < -1 and breadth < -1000:
        return "弱势"
    if avg_index < -0.25 and breadth < 0:
        return "震荡偏弱"
    return "震荡"


def emotion_score(up_count, down_count, limit_up, limit_down):
    breadth_score = max(0, min(4, (up_count - down_count + 3000) / 1500))
    limit_score = max(0, min(4, limit_up / 25))
    risk_penalty = max(0, min(3, limit_down / 12))
    return round(max(1, min(10, breadth_score + limit_score + 3 - risk_penalty)), 1)


def build_review():
    now = datetime.now()
    source_notes = []
    try:
        indices = get_indices()
    except Exception as exc:
        log(f"all index sources failed, using cached indices: {exc}")
        indices = get_cached_field("indices", [])
        source_notes.append("indices_cached")
        if not indices:
            raise
    stocks = get_a_shares()
    boards = get_boards()
    if not stocks:
        raise RuntimeError("未获取到A股列表，可能是网络或行情接口暂不可用。")

    up_count = sum(1 for s in stocks if s["change_pct"] > 0)
    down_count = sum(1 for s in stocks if s["change_pct"] < 0)
    flat_count = len(stocks) - up_count - down_count
    limit_up = sum(1 for s in stocks if is_limit_up(s) and not is_st_stock(s))
    limit_down = sum(1 for s in stocks if s["change_pct"] <= -9.8)
    turnover_yi = yuan_to_yi(sum(s["amount_yuan"] for s in stocks))
    if len(stocks) < 500:
        # Sina rank fallback is a focused sample rather than the full market.
        # Use 上证 + 深成指成交额 as the broad market proxy; 创业板已包含在深市成交中.
        turnover_yi = yuan_to_yi(sum(i["amount_yuan"] for i in indices[:2]))
        up_count = None
        down_count = None
        flat_count = None
    main_net_yi = yuan_to_yi(sum(s["main_net_yuan"] for s in stocks))
    top_stocks = sorted(stocks, key=lambda x: (x["change_pct"], x["amount_yuan"]), reverse=True)[:20]
    limit_up_stocks = build_limit_up_stocks(stocks)
    history_before = load_history()
    apply_stock_tracking(limit_up_stocks, history_before)
    concept_boards = build_concept_boards(limit_up_stocks)
    apply_board_reasons(concept_boards)
    top_boards = sorted(boards, key=lambda x: (x["change_pct"], x["main_net_yuan"]), reverse=True)[:12]
    if not top_boards:
        top_boards = [
            {
                "code": "",
                "name": item["concept"],
                "change_pct": item["avg_change_pct"],
                "main_net_yuan": 0,
                "leader": item["leader"],
                "limit_up_count": item["limit_up_count"],
            }
            for item in concept_boards[:12]
        ] or infer_themes_from_stocks(top_stocks)
    inflow_boards = sorted(boards, key=lambda x: x["main_net_yuan"], reverse=True)[:12]
    if up_count is None:
        avg_index = sum(i["change_pct"] for i in indices[:3]) / max(len(indices[:3]), 1)
        state = "震荡偏强" if avg_index > 0.25 else "震荡" if avg_index > -0.25 else "震荡偏弱"
        score = 6.5 if state == "震荡偏强" else 5.0
    else:
        state = market_state(indices, up_count, down_count, limit_up, limit_down)
        score = emotion_score(up_count, down_count, limit_up, limit_down)
    strongest = top_boards[0]["name"] if top_boards else ""

    conclusion = (
        f"市场{state}，指数分化但创业板较强，两市成交额约{turnover_yi}亿元，"
        f"{'上涨'+str(up_count)+'家、下跌'+str(down_count)+'家，' if up_count is not None else '个股宽度使用排行样本估算，'}"
        f"强势样本中涨停/20cm个股活跃。"
        f"当前强势方向：{strongest or '待确认'}。"
    )

    payload = {
        "updated_at": now.strftime("%Y-%m-%d %H:%M:%S"),
        "data_updated_at": now.strftime("%Y-%m-%d %H:%M:%S"),
        "last_checked_at": now.strftime("%Y-%m-%d %H:%M:%S"),
        "trade_date": now.strftime("%Y-%m-%d"),
        "status": "ok",
        "source": "Tencent index quote + Sina A-share rank fallback; Eastmoney when available",
        "source_status": {
            "mode": "fast_fallback" if USE_FAST_FALLBACK else "eastmoney_primary",
            "indices": "Tencent qt.gtimg.cn",
            "stocks": "Sina A-share rank top 180 by changepercent",
            "concepts": "Local concept rules from涨停股名称匹配",
            "full_market_breadth": "unavailable_in_fast_fallback",
            "fund_flow": "unavailable_in_fast_fallback",
            "detail_fundamentals": "partial_from_sina_rank",
            "freshness": "fresh_fetch",
            "notes": source_notes,
            "updated_at": now.strftime("%Y-%m-%d %H:%M:%S"),
        },
        "indices": indices,
        "market": {
            "turnover_yi": turnover_yi,
            "main_net_yi": main_net_yi,
            "up_count": up_count,
            "down_count": down_count,
            "flat_count": flat_count,
            "limit_up": limit_up,
            "limit_down": limit_down,
            "state": state,
            "emotion_score": score,
            "suggested_position": position_by_state(state, score),
            "conclusion": conclusion,
        },
        "themes": top_boards,
        "concept_boards": concept_boards,
        "limit_up_stocks": limit_up_stocks,
        "hot_board_tracking": build_hot_board_tracking(concept_boards, history_before),
        "hot_stock_tracking": build_hot_stock_tracking(
            concept_boards,
            limit_up_stocks,
            history_before,
            now.strftime("%Y-%m-%d"),
        ),
        "fund_inflow_themes": inflow_boards,
        "core_stocks": top_stocks,
        "plan": make_plan(state, score, top_boards),
    }
    save_history_snapshot(payload)
    return payload


def is_limit_up(stock):
    code = str(stock.get("code") or "")
    change = stock.get("change_pct", 0)
    if code.startswith(("300", "301", "688", "920", "430", "830", "831", "832", "833", "834", "835", "836", "837", "838", "839")):
        return change >= 19.8
    if "ST" in (stock.get("name") or "").upper():
        return change >= 4.8
    return change >= 9.8


def is_st_stock(stock):
    return "ST" in (stock.get("name") or "").upper()


def concepts_for_stock(stock):
    name = stock.get("name") or ""
    concepts = []
    for concept, keys in CONCEPT_RULES:
        if any(key and key in name for key in keys):
            concepts.append(concept)
    if not concepts:
        concepts.append("其他涨停")
    return concepts


def build_limit_up_stocks(stocks):
    result = []
    for stock in stocks:
        if is_st_stock(stock):
            continue
        if not is_limit_up(stock):
            continue
        item = dict(stock)
        item["concepts"] = concepts_for_stock(stock)
        item["limit_reason"] = reason_for_stock(item)
        result.append(item)
    return sorted(result, key=lambda x: (x["change_pct"], x["amount_yuan"]), reverse=True)


def build_concept_boards(limit_up_stocks):
    grouped = {}
    for stock in limit_up_stocks:
        for concept in stock["concepts"]:
            grouped.setdefault(concept, []).append(stock)

    raw_boards = []
    for concept, members in grouped.items():
        ordered = sorted(members, key=lambda x: (x["change_pct"], x["amount_yuan"]), reverse=True)
        raw_boards.append(
            {
                "concept": concept,
                "limit_up_count": len(ordered),
                "leader": ordered[0]["name"],
                "avg_change_pct": round(sum(x["change_pct"] for x in ordered) / len(ordered), 2),
                "stocks": [
                    {
                        "code": s["code"],
                        "name": s["name"],
                        "price": s.get("price"),
                        "change_pct": s["change_pct"],
                        "price_change": s.get("price_change"),
                        "pre_close": s.get("pre_close"),
                        "open": s.get("open"),
                        "high": s.get("high"),
                        "low": s.get("low"),
                        "volume": s.get("volume"),
                        "amount_yuan": s.get("amount_yuan"),
                        "turnover_pct": s["turnover_pct"],
                        "pe": s.get("pe"),
                        "pb": s.get("pb"),
                        "market_cap_wan": s.get("market_cap_wan"),
                        "float_market_cap_wan": s.get("float_market_cap_wan"),
                        "ticktime": s.get("ticktime"),
                        "source": s.get("source"),
                        "concepts": s["concepts"],
                        "limit_reason": s.get("limit_reason") or reason_for_stock(s),
                        "new_concepts": s.get("new_concepts", []),
                    }
                    for s in ordered
                ],
            }
        )

    boards = [item for item in raw_boards if item["concept"] != "其他涨停" and item["limit_up_count"] >= 2]
    retained_stock_codes = {
        stock["code"]
        for board in boards
        for stock in board.get("stocks", [])
    }
    merged_other = {}
    for item in raw_boards:
        if item["concept"] == "其他涨停" or item["limit_up_count"] == 1:
            for stock in item["stocks"]:
                if stock["code"] not in retained_stock_codes:
                    merged_other[stock["code"]] = stock

    if merged_other:
        other_stocks = sorted(
            merged_other.values(),
            key=lambda x: (x["change_pct"], x.get("amount_yuan", 0)),
            reverse=True,
        )
        boards.append(
            {
                "concept": "其他涨停",
                "limit_up_count": len(other_stocks),
                "leader": other_stocks[0]["name"],
                "avg_change_pct": round(sum(x["change_pct"] for x in other_stocks) / len(other_stocks), 2),
                "reason": CONCEPT_REASON["其他涨停"],
                "stocks": other_stocks,
            }
        )
    return sorted(
        boards,
        key=lambda x: (
            x["concept"] != "其他涨停",
            x["limit_up_count"],
            x["avg_change_pct"],
        ),
        reverse=True,
    )


def reason_for_stock(stock):
    concepts = stock.get("concepts") or ["其他涨停"]
    main = concepts[0]
    reason = CONCEPT_REASON.get(main, CONCEPT_REASON["其他涨停"])
    if len(concepts) > 1:
        return f"{reason}；叠加{'、'.join(concepts[1:3])}概念"
    return reason


def apply_board_reasons(boards):
    for board in boards:
        board["reason"] = CONCEPT_REASON.get(board["concept"], CONCEPT_REASON["其他涨停"])


def load_history():
    if not HISTORY_FILE.exists():
        return []
    try:
        return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def save_history_snapshot(payload):
    history = load_history()
    trade_date = payload.get("trade_date")
    snapshot = {
        "trade_date": trade_date,
        "updated_at": payload.get("data_updated_at") or payload.get("updated_at"),
        "concept_boards": [
            {
                "concept": b["concept"],
                "limit_up_count": b["limit_up_count"],
                "leader": b["leader"],
                "stocks": [s["name"] for s in b.get("stocks", [])],
            }
            for b in payload.get("concept_boards", [])
        ],
        "limit_up_stocks": [
            {
                "code": s["code"],
                "name": s["name"],
                "price": s.get("price"),
                "change_pct": s.get("change_pct"),
                "amount_yuan": s.get("amount_yuan"),
                "turnover_pct": s.get("turnover_pct"),
                "concepts": s.get("concepts", []),
                "limit_reason": s.get("limit_reason"),
            }
            for s in payload.get("limit_up_stocks", [])
        ],
    }
    history = [item for item in history if item.get("trade_date") != trade_date]
    history.append(snapshot)
    history = sorted(history, key=lambda x: x.get("trade_date", ""))[-10:]
    HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def apply_stock_tracking(limit_up_stocks, history):
    previous_by_code = {}
    for day in history[-5:]:
        for stock in day.get("limit_up_stocks", []):
            previous_by_code[stock.get("code")] = set(stock.get("concepts", []))
    for stock in limit_up_stocks:
        prev = previous_by_code.get(stock.get("code"), set())
        current = set(stock.get("concepts", []))
        stock["new_concepts"] = sorted(current - prev) if prev else stock.get("concepts", [])


def build_hot_board_tracking(concept_boards, history):
    recent = history[-2:] + [
        {
            "trade_date": datetime.now().strftime("%Y-%m-%d"),
            "concept_boards": [
                {
                    "concept": b["concept"],
                    "limit_up_count": b["limit_up_count"],
                    "leader": b["leader"],
                    "stocks": [s["name"] for s in b.get("stocks", [])],
                }
                for b in concept_boards
            ],
        }
    ]
    concept_map = {}
    for day in recent:
        for board in day.get("concept_boards", []):
            concept = board["concept"]
            concept_map.setdefault(concept, []).append(
                {
                    "date": day.get("trade_date"),
                    "count": board.get("limit_up_count", 0),
                    "leader": board.get("leader", ""),
                    "stocks": board.get("stocks", []),
                }
            )
    rows = []
    for concept, days in concept_map.items():
        active_days = [d for d in days if d["count"] >= 2]
        rows.append(
            {
                "concept": concept,
                "active_days": len(active_days),
                "latest_count": days[-1]["count"],
                "trend": "连续活跃" if len(active_days) >= 2 else "今日活跃",
                "highlight": len(active_days) >= 2,
                "history": days,
            }
        )
    return sorted(rows, key=lambda x: (x["highlight"], x["active_days"], x["latest_count"]), reverse=True)[:12]


def build_hot_stock_tracking(concept_boards, limit_up_stocks, history, trade_date):
    board_strength = {}
    for board in concept_boards:
        concept = board["concept"]
        previous_days = []
        for day in history[-2:]:
            matched = next((b for b in day.get("concept_boards", []) if b.get("concept") == concept), None)
            if matched and matched.get("limit_up_count", 0) >= 2:
                previous_days.append(day.get("trade_date"))
        board_strength[concept] = {
            "leader": board.get("leader"),
            "count": board.get("limit_up_count", 0),
            "active_days": len(previous_days) + (1 if board.get("limit_up_count", 0) >= 2 else 0),
            "previous_days": previous_days,
        }

    previous_by_code = {}
    for day in history[-3:]:
        for stock in day.get("limit_up_stocks", []):
            previous_by_code.setdefault(stock.get("code"), []).append(day.get("trade_date"))

    candidates = {}
    for board in concept_boards:
        concept = board["concept"]
        if board.get("limit_up_count", 0) < 2 and concept != "其他涨停":
            continue
        for index, stock in enumerate((board.get("stocks") or [])[:4]):
            key = stock.get("code") or stock.get("name")
            if not key:
                continue
            current = candidates.get(key)
            role_score = 3 if stock.get("name") == board.get("leader") else max(1, 3 - index)
            score = board.get("limit_up_count", 0) * 10 + role_score + int(stock.get("change_pct", 0))
            if current and current["score"] >= score:
                continue

            previous_dates = previous_by_code.get(stock.get("code"), [])
            features = []
            if stock.get("name") == board.get("leader"):
                features.append("板块龙头")
            elif index <= 2:
                features.append("前排领涨")
            if stock.get("change_pct", 0) >= 19:
                features.append("20cm涨停")
            elif "ST" in stock.get("name", ""):
                features.append("ST修复")
            else:
                features.append("涨停确认")
            if stock.get("turnover_pct", 0) >= 15:
                features.append("放量分歧")
            if stock.get("new_concepts"):
                features.append("新增概念")
            if board_strength.get(concept, {}).get("active_days", 0) >= 2:
                features.append("板块连续")

            active_days = board_strength.get(concept, {}).get("active_days", 1)
            if previous_dates:
                consensus_day = f"{previous_dates[-1]} + 延续"
                divergence_day = trade_date if stock.get("turnover_pct", 0) >= 12 else "延续观察"
            else:
                consensus_day = f"{trade_date} + 首板"
                divergence_day = trade_date

            if stock.get("turnover_pct", 0) >= 15:
                tomorrow_view = "分歧放量，重点看开盘承接和回封力度"
            elif stock.get("name") == board.get("leader") and active_days >= 2:
                tomorrow_view = "板块龙头，弱转强或继续封板则主线延续"
            elif stock.get("new_concepts"):
                tomorrow_view = "新增概念发酵，观察同概念是否继续扩散"
            elif concept == "其他涨停":
                tomorrow_view = "先看是否被市场重新归入明确主线"
            else:
                tomorrow_view = "看所属板块涨停家数是否增加，个股是否继续前排"

            candidates[key] = {
                "score": score,
                "code": stock.get("code"),
                "name": stock.get("name"),
                "concept": concept,
                "price": stock.get("price"),
                "change_pct": stock.get("change_pct"),
                "divergence_day": divergence_day,
                "features": features[:4],
                "consensus_day": consensus_day,
                "tomorrow_view": tomorrow_view,
                "highlight": stock.get("name") == board.get("leader") or active_days >= 2 or stock.get("change_pct", 0) >= 19,
            }

    return [
        {k: v for k, v in item.items() if k != "score"}
        for item in sorted(candidates.values(), key=lambda x: (x["highlight"], x["score"]), reverse=True)[:18]
    ]


def infer_themes_from_stocks(stocks):
    rules = [
        ("CPO/光通信", ["太辰光", "光库科技", "铭普光磁", "通鼎互联", "长进光子"]),
        ("半导体/PCB/电子", ["南芯科技", "雅创电子", "本川智能", "华曙高科", "中巨芯", "大普微", "康强电子"]),
        ("机器人/智能制造", ["先导智能", "科大智能", "乔锋智能", "德恩精工", "安达智能"]),
        ("新材料/复合材料", ["金戈新材", "逸豪新材", "国际复材", "锦华新材", "同宇新材"]),
        ("锂电/新能源", ["亿纬锂能", "珠海冠宇", "多氟多", "盛新锂能", "天山铝业"]),
    ]
    themes = []
    for name, keys in rules:
        matched = [s for s in stocks if any(k in s["name"] for k in keys)]
        if matched:
            leader = max(matched, key=lambda x: x["change_pct"])
            themes.append(
                {
                    "code": "",
                    "name": name,
                    "change_pct": round(sum(x["change_pct"] for x in matched) / len(matched), 2),
                    "main_net_yuan": 0,
                    "leader": leader["name"],
                }
            )
    return themes[:12]


def position_by_state(state, score):
    if state == "强势" and score >= 7:
        return "5-7成"
    if state in ("震荡偏强", "强势"):
        return "3-5成"
    if state == "震荡":
        return "2-4成"
    if state == "震荡偏弱":
        return "1-2成"
    return "空仓或1成试错"


def make_plan(state, score, themes):
    names = [x["name"] for x in themes[:3]]
    if state in ("强势", "震荡偏强"):
        trigger = "指数不明显放量下杀，核心方向前排继续强于大盘。"
    elif state == "震荡":
        trigger = "成交额不继续萎缩，热点有明确核心股带动。"
    else:
        trigger = "先等指数止跌、跌停数下降、高位股负反馈缓和。"
    return {
        "watch_themes": names,
        "trigger": trigger,
        "invalid": "成交额明显萎缩、热点前排冲高回落、跌停或大面个股增加。",
        "risk": "自动生成仅用于复盘观察，不构成投资建议；交易前需要结合个人仓位和风险承受能力。",
        "position": position_by_state(state, score),
    }


def save_json(payload):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    tmp = DATA_FILE.with_suffix(".json.tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(DATA_FILE)
    DATA_JS_FILE.write_text(f"window.__REVIEW_DATA__ = {text};", encoding="utf-8")


def get_cached_field(field, default=None):
    if not DATA_FILE.exists():
        return default
    try:
        payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
        return payload.get(field, default)
    except Exception:
        return default


def rebuild_concepts_from_existing():
    if not DATA_FILE.exists():
        raise RuntimeError("没有可复用的历史数据文件。")
    payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    if payload.get("status") != "ok":
        raise RuntimeError("历史数据不是成功状态，无法复用。")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payload.setdefault("data_updated_at", payload.get("updated_at"))
    stocks = payload.get("limit_up_stocks") or payload.get("core_stocks") or []
    limit_up_stocks = build_limit_up_stocks(stocks)
    history_before = load_history()
    apply_stock_tracking(limit_up_stocks, history_before)
    concept_boards = build_concept_boards(limit_up_stocks)
    payload["concept_boards"] = concept_boards
    payload["limit_up_stocks"] = limit_up_stocks
    payload["hot_board_tracking"] = build_hot_board_tracking(concept_boards, history_before)
    payload["hot_stock_tracking"] = build_hot_stock_tracking(
        concept_boards,
        limit_up_stocks,
        history_before,
        payload.get("trade_date") or now[:10],
    )
    if concept_boards:
        payload["themes"] = [
            {
                "code": "",
                "name": item["concept"],
                "change_pct": item["avg_change_pct"],
                "main_net_yuan": 0,
                "leader": item["leader"],
                "limit_up_count": item["limit_up_count"],
            }
            for item in concept_boards[:12]
        ]
        payload.setdefault("plan", {})["watch_themes"] = [x["concept"] for x in concept_boards[:3]]
    payload["last_checked_at"] = now
    payload["source_status"] = {
        "mode": "reused_existing_after_fetch_failed",
        "indices": "cached from previous successful refresh",
        "stocks": "cached from previous successful refresh",
        "concepts": "Local concept rules recomputed from cached涨停股",
        "full_market_breadth": "unavailable_in_fast_fallback",
        "fund_flow": "unavailable_in_fast_fallback",
        "detail_fundamentals": "partial_from_cached_sina_rank",
        "freshness": "cached_data_recomputed",
        "updated_at": now,
    }
    save_json(payload)
    return payload


def update_workbook(payload):
    if not WORKBOOK_FILE.exists():
        return
    wb = load_workbook(WORKBOOK_FILE)
    ws = wb["每日复盘"]
    trade_date = payload["trade_date"]
    target = None
    for row in range(3, 203):
        value = ws[f"A{row}"].value
        if value in (None, ""):
            target = row
            break
        if str(value)[:10] == trade_date:
            target = row
            break
    if target is None:
        target = 202

    indices = {i["code"]: i for i in payload["indices"]}
    market = payload["market"]
    ws[f"A{target}"] = trade_date
    ws[f"B{target}"] = indices.get("000001", {}).get("change_pct", 0) / 100
    ws[f"C{target}"] = indices.get("399001", {}).get("change_pct", 0) / 100
    ws[f"D{target}"] = indices.get("399006", {}).get("change_pct", 0) / 100
    ws[f"E{target}"] = market["turnover_yi"]
    ws[f"G{target}"] = market["up_count"]
    ws[f"H{target}"] = market["state"]
    ws[f"I{target}"] = market["limit_up"]
    ws[f"J{target}"] = market["limit_down"]
    ws[f"K{target}"] = market["main_net_yi"]
    ws[f"L{target}"] = market["emotion_score"]
    ws[f"M{target}"] = market["suggested_position"]
    ws[f"N{target}"] = "、".join(payload["plan"]["watch_themes"])
    ws[f"O{target}"] = "、".join(t["name"] for t in payload["fund_inflow_themes"][:5])
    ws[f"Q{target}"] = market["conclusion"]
    wb.save(WORKBOOK_FILE)


def main():
    try:
        payload = build_review()
        save_json(payload)
        update_workbook(payload)
        log(f"refresh ok: {payload['trade_date']}")
        print(json.dumps({"status": "ok", "file": str(DATA_FILE)}, ensure_ascii=False))
    except Exception as exc:
        try:
            payload = rebuild_concepts_from_existing()
            update_workbook(payload)
            log(f"refresh reused existing data after fetch failed: {exc}")
            print(json.dumps({"status": "ok", "file": str(DATA_FILE), "reused_existing": True}, ensure_ascii=False))
            return 0
        except Exception as reuse_exc:
            log(f"reuse existing data failed: {reuse_exc}")
        fallback = {
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "trade_date": datetime.now().strftime("%Y-%m-%d"),
            "status": "error",
            "error": str(exc),
        }
        if not DATA_FILE.exists():
            save_json(fallback)
        log(f"refresh failed: {exc}")
        print(json.dumps(fallback, ensure_ascii=False))
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
