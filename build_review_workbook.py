from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT_DIR = Path(r"C:\Users\zeng\Documents\万户侯\outputs\a_share_daily_review")
OUT_FILE = OUT_DIR / "每日A股复盘表.xlsx"

COLORS = {
    "dark": "23395B",
    "teal": "1F7A8C",
    "light": "E8F1F2",
    "line": "D9E2EC",
    "white": "FFFFFF",
}


def style_range(ws, cell_range, fill=None, font=None, align=None, border=True):
    thin = Side(style="thin", color=COLORS["line"])
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if border:
                cell.border = box


def title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.fill = PatternFill("solid", fgColor=COLORS["dark"])
    c.font = Font(bold=True, color=COLORS["white"], size=15)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[c.row].height = 26
    ws.sheet_view.showGridLines = False


def header(ws, row, last_col):
    fill = PatternFill("solid", fgColor=COLORS["teal"])
    font = Font(bold=True, color=COLORS["white"])
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    style_range(ws, f"A{row}:{get_column_letter(last_col)}{row}", fill, font, align)
    ws.row_dimensions[row].height = 28


def body(ws, cell_range):
    style_range(
        ws,
        cell_range,
        PatternFill("solid", fgColor=COLORS["white"]),
        None,
        Alignment(vertical="top", wrap_text=True),
    )


def widths(ws, values):
    for idx, width in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_list(ws, cell_range, values):
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    dash = wb.create_sheet("复盘仪表盘")
    daily = wb.create_sheet("每日复盘")
    themes = wb.create_sheet("板块热点")
    stocks = wb.create_sheet("核心个股")
    positions = wb.create_sheet("持仓计划")
    rules = wb.create_sheet("复盘规则")

    title(dash, "A1:J1", "每日A股复盘仪表盘")
    dash["A3"] = "今日市场结论：填写市场强弱、资金方向、明日仓位。"
    dash.merge_cells("A3:J3")
    body(dash, "A3:J3")
    kpis = [
        ("最新日期", '=IFERROR(MAX(\'每日复盘\'!A3:A202),"")'),
        ("市场状态", '=IFERROR(LOOKUP(2,1/(\'每日复盘\'!A3:A202<>""),\'每日复盘\'!H3:H202),"")'),
        ("成交额(亿元)", '=IFERROR(LOOKUP(2,1/(\'每日复盘\'!A3:A202<>""),\'每日复盘\'!E3:E202),"")'),
        ("情绪评分", '=IFERROR(LOOKUP(2,1/(\'每日复盘\'!A3:A202<>""),\'每日复盘\'!L3:L202),"")'),
        ("建议仓位", '=IFERROR(LOOKUP(2,1/(\'每日复盘\'!A3:A202<>""),\'每日复盘\'!M3:M202),"")'),
    ]
    for r, (label, formula) in enumerate(kpis, 5):
        dash[f"A{r}"] = label
        dash[f"B{r}"] = formula
    style_range(dash, "A5:A9", PatternFill("solid", fgColor=COLORS["light"]), Font(bold=True, color=COLORS["dark"]))
    body(dash, "B5:B9")
    dash["D5"] = "日期"
    dash["E5"] = "成交额"
    dash["F5"] = "涨停"
    dash["G5"] = "跌停"
    dash["H5"] = "最强方向"
    dash["I5"] = "市场状态"
    dash["J5"] = "仓位"
    header(dash, 5, 10)
    body(dash, "D6:J10")
    dash["A12"] = "明日交易计划"
    dash.merge_cells("A12:J12")
    style_range(dash, "A12:J12", PatternFill("solid", fgColor=COLORS["light"]), Font(bold=True, color=COLORS["dark"]), border=False)
    plan_rows = [
        ["重点方向1", "", "观察股", "", "触发条件", "", "失效条件", "", "计划仓位", ""],
        ["重点方向2", "", "观察股", "", "触发条件", "", "失效条件", "", "计划仓位", ""],
        ["重点方向3", "", "观察股", "", "触发条件", "", "失效条件", "", "计划仓位", ""],
        ["风险提醒", "", "", "", "", "", "", "", "", ""],
        ["盘前确认", "", "指数/量能/高位股反馈/消息面/外围市场", "", "", "", "", "", "", ""],
    ]
    for r, row in enumerate(plan_rows, 13):
        for c, value in enumerate(row, 1):
            dash.cell(r, c, value)
    body(dash, "A13:J17")
    widths(dash, [14, 15, 10, 14, 12, 20, 12, 20, 12, 12])

    daily_headers = ["日期", "上证%", "深成%", "创业板%", "两市成交额(亿元)", "较昨日增减(亿元)", "上涨家数", "市场状态", "涨停数", "跌停数", "北向/替代资金(亿元)", "情绪评分(1-10)", "建议仓位", "最强方向", "资金流入方向", "资金流出方向", "一句话复盘"]
    title(daily, "A1:Q1", "每日复盘记录")
    daily.append(daily_headers)
    header(daily, 2, len(daily_headers))
    body(daily, "A3:Q202")
    add_list(daily, "H3:H202", ["强势", "震荡偏强", "震荡", "震荡偏弱", "弱势"])
    add_list(daily, "M3:M202", ["空仓", "1成", "2成", "3成", "4成", "5成", "6成", "7成", "8成"])
    daily.freeze_panes = "A3"
    widths(daily, [12, 9, 9, 9, 15, 15, 10, 12, 9, 9, 16, 12, 10, 16, 20, 20, 32])

    theme_headers = ["日期", "板块/概念", "类型", "板块涨幅%", "资金净流入(亿元)", "涨停数", "核心龙头", "梯队/扩散", "催化逻辑", "持续性(1-5)", "强度评分", "明日策略", "观察条件", "风险点", "备注"]
    title(themes, "A1:O1", "板块热点与资金跟踪")
    themes.append(theme_headers)
    header(themes, 2, len(theme_headers))
    body(themes, "A3:O152")
    for r in range(3, 153):
        themes[f"K{r}"] = f'=IF(B{r}="","",J{r}*2+MIN(F{r},5)+IF(C{r}="主线",3,IF(C{r}="轮动",1,0)))'
    add_list(themes, "C3:C152", ["主线", "轮动", "消息刺激", "防守避险", "退潮"])
    add_list(themes, "J3:J152", ["1", "2", "3", "4", "5"])
    add_list(themes, "L3:L152", ["重点跟踪", "低吸观察", "只看不追", "回避"])
    themes.freeze_panes = "A3"
    widths(themes, [12, 16, 12, 12, 16, 9, 16, 16, 28, 12, 11, 12, 24, 20, 20])

    stock_headers = ["日期", "代码", "名称", "所属板块", "角色", "收盘表现", "量能", "位置", "换手率%", "是否涨停", "买入触发", "失效条件", "预期空间", "跟踪结论", "次日动作", "备注"]
    title(stocks, "A1:P1", "核心个股观察池")
    stocks.append(stock_headers)
    header(stocks, 2, len(stock_headers))
    body(stocks, "A3:P152")
    add_list(stocks, "E3:E152", ["龙头", "中军", "补涨", "趋势核心", "套利", "风险观察"])
    add_list(stocks, "G3:G152", ["放量", "缩量", "温和放量", "巨量分歧", "量能不足"])
    add_list(stocks, "H3:H152", ["低位启动", "突破位", "趋势中继", "高位加速", "高位分歧", "破位"])
    add_list(stocks, "J3:J152", ["是", "否"])
    add_list(stocks, "O3:O152", ["买入观察", "持有观察", "减仓观察", "放弃", "等待确认"])
    stocks.freeze_panes = "A3"
    widths(stocks, [12, 10, 12, 16, 12, 16, 12, 14, 11, 11, 24, 24, 14, 20, 12, 20])

    position_headers = ["日期", "代码", "名称", "方向", "仓位%", "成本价", "现价", "浮盈亏%", "止损价", "止盈/目标价", "买入理由", "今日处理", "明日计划", "是否执行", "复盘评分(1-5)", "备注"]
    title(positions, "A1:P1", "持仓与交易计划")
    positions.append(position_headers)
    header(positions, 2, len(position_headers))
    body(positions, "A3:P102")
    for r in range(3, 103):
        positions[f"H{r}"] = f'=IF(OR(F{r}="",G{r}=""),"",G{r}/F{r}-1)'
    add_list(positions, "D3:D102", ["短线", "波段", "趋势", "防守", "观察"])
    add_list(positions, "N3:N102", ["已执行", "部分执行", "未执行", "取消"])
    add_list(positions, "O3:O102", ["1", "2", "3", "4", "5"])
    positions.freeze_panes = "A3"
    widths(positions, [12, 10, 12, 10, 9, 9, 9, 10, 9, 12, 24, 16, 22, 12, 14, 18])

    title(rules, "A1:H1", "复盘规则与打分标准")
    rules.append([])
    rules.append(["模块", "检查项", "判断标准", "强信号", "弱信号", "行动建议", "记录位置", "备注"])
    header(rules, 3, 8)
    rule_rows = [
        ["市场环境", "量能", "成交额较昨日变化", "放量上涨或缩量止跌", "缩量上涨或放量下跌", "决定基础仓位", "每日复盘", "先判环境再选股"],
        ["赚钱效应", "涨跌家数/涨跌停", "上涨家数、涨停数、跌停数", "涨停扩散且跌停少", "跌停增加、高位股负反馈", "控制仓位", "每日复盘", "情绪评分1-10"],
        ["热点识别", "板块持续性", "连续活跃天数与核心股反馈", "主线延续、有梯队", "一日游、无核心", "只做主线或低吸", "板块热点", "避免只看涨幅榜"],
        ["核心个股", "角色定位", "龙头/中军/补涨/套利", "龙头强，中军稳", "龙头断板且补涨混乱", "优先核心，少做杂股", "核心个股", "记录失效条件"],
        ["交易计划", "触发条件", "指数、板块、个股三者共振", "放量突破/回踩确认", "冲高无量、跌破关键位", "计划内交易", "复盘仪表盘", "不因盘中情绪改规则"],
        ["持仓管理", "止损止盈", "价格/逻辑双止损", "未破位且逻辑延续", "跌破止损或逻辑证伪", "执行计划", "持仓计划", "复盘是否执行"],
    ]
    for row in rule_rows:
        rules.append(row)
    body(rules, "A4:H9")
    rules["A12"] = "每日复盘顺序：1 指数和成交额  2 赚钱效应  3 资金方向  4 热点持续性  5 明日计划  6 次日复盘"
    rules.merge_cells("A12:H12")
    body(rules, "A12:H12")
    widths(rules, [12, 16, 24, 22, 22, 20, 14, 22])

    for ws in [daily, themes, stocks, positions]:
        ws.auto_filter.ref = ws.dimensions
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 1 and cell.row >= 3:
                    cell.number_format = "yyyy-mm-dd"
        ws.sheet_view.zoomScale = 90

    wb.save(OUT_FILE)

    check = load_workbook(OUT_FILE, data_only=False)
    assert set(check.sheetnames) == {"复盘仪表盘", "每日复盘", "板块热点", "核心个股", "持仓计划", "复盘规则"}
    assert check["持仓计划"]["H3"].value.startswith("=IF")
    assert check["板块热点"]["K3"].value.startswith("=IF")
    print(OUT_FILE)


if __name__ == "__main__":
    main()
